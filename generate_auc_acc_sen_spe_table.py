import numpy as np
import pandas as pd
import os
import h5py
from sklearn import metrics
from matplotlib import pyplot as plt
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, roc_auc_score, roc_curve, auc,accuracy_score
import openpyxl
from openpyxl.styles import Font, Alignment

def load_csv_file(filename1,filename2,dir1, dir2,filename3=None):
    
    
    if filename3:
        file_path_dir31 = os.path.join(dir1, filename3)
        file_path_dir32 = os.path.join(dir2, filename3)
        file_path_dir1 = os.path.join(dir1, filename1)
        file_path_dir2 = os.path.join(dir2, filename2)
        if os.path.isfile(file_path_dir31) and os.path.isfile(file_path_dir1):
            print(f"File {filename3} found in {dir1}.")
            return file_path_dir1,file_path_dir31
        elif os.path.isfile(file_path_dir32) and os.path.isfile(file_path_dir2):
            print(f"File not found in {dir1}. Checking {dir2}...")
            print(f"File {filename3} found in {dir2}.")
            return file_path_dir2,file_path_dir32
        else:
            print("File not found in both directories.")
            return None
    else:
        file_path_dir1 = os.path.join(dir1, filename1)
        file_path_dir2 = os.path.join(dir2, filename2)
        if os.path.isfile(file_path_dir1):
            print(f"File {filename1} found in {dir1}.")
            return file_path_dir1
        elif os.path.isfile(file_path_dir2):
            print(f"File not found in {dir1}. Checking {dir2}...")
            print(f"File {filename2} found in {dir2}.")
            return file_path_dir2
        else:
            print("File not found in both directories.")
            return None

def create_excel(mit_m,longfor_m,resnet_m,espformer_m):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    
    # Rename the worksheet (optional)
    # sheet.title = "Model Performance"
    
    bold_font = Font(bold=True)
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    
    
    sheet['A1'] = 'BM'
    sheet['A1'].font = bold_font
    sheet['A1'].alignment = center_alignment
    
    
    models = ['ViT', 'ESPFormer', 'LLM', 'ResNet']
    start_col = 2
    
    for model in models:
        
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+3)
        sheet.cell(row=1, column=start_col).value = model
        sheet.cell(row=1, column=start_col).font = bold_font
        sheet.cell(row=1, column=start_col).alignment = center_alignment
        
        # Sub-columns for Acc., Sen., Spe.

        sheet.cell(row=2, column=start_col).value = 'AUC.'
        sheet.cell(row=2, column=start_col).font = bold_font
        sheet.cell(row=2, column=start_col).alignment = center_alignment

        sheet.cell(row=2, column=start_col+1).value = 'Acc.'
        sheet.cell(row=2, column=start_col+1).font = bold_font
        sheet.cell(row=2, column=start_col+1).alignment = center_alignment
    
        sheet.cell(row=2, column=start_col+2).value = 'Sen.'
        sheet.cell(row=2, column=start_col+2).font = bold_font
        sheet.cell(row=2, column=start_col+2).alignment = center_alignment
    
        sheet.cell(row=2, column=start_col+3).value = 'Spe.'
        sheet.cell(row=2, column=start_col+3).font = bold_font
        sheet.cell(row=2, column=start_col+3).alignment = center_alignment


        
        
        
        start_col += 4
    

    for i in range(1, 13):
        sheet[f'A{i+2}'] = i
        sheet[f'A{i+2}'].alignment = center_alignment
        # sheet[f'A{i+2}'].border = thin_border
        
        sheet[f'B{i+2}'] = mit_m[i-1][0]
        sheet[f'B{i+2}'].number_format = '0.00%'
        sheet[f'C{i+2}'] = mit_m[i-1][1]
        sheet[f'C{i+2}'].number_format = '0.00%'
        sheet[f'D{i+2}'] = mit_m[i-1][2]
        sheet[f'D{i+2}'].number_format = '0.00%'
        sheet[f'E{i+2}'] = mit_m[i-1][3]
        sheet[f'E{i+2}'].number_format = '0.00%'

        sheet[f'F{i+2}'] = espformer_m[i-1][0]
        sheet[f'F{i+2}'].number_format = '0.00%'
        sheet[f'G{i+2}'] = espformer_m[i-1][1]
        sheet[f'G{i+2}'].number_format = '0.00%'
        sheet[f'H{i+2}'] = espformer_m[i-1][2]
        sheet[f'H{i+2}'].number_format = '0.00%'
        sheet[f'I{i+2}'] = espformer_m[i-1][3]
        sheet[f'I{i+2}'].number_format = '0.00%'

        sheet[f'J{i+2}'] = longfor_m[i-1][0]
        sheet[f'J{i+2}'].number_format = '0.00%'
        sheet[f'K{i+2}'] = longfor_m[i-1][1]
        sheet[f'K{i+2}'].number_format = '0.00%'
        sheet[f'L{i+2}'] = longfor_m[i-1][2]
        sheet[f'L{i+2}'].number_format = '0.00%'
        sheet[f'M{i+2}'] = longfor_m[i-1][3]
        sheet[f'M{i+2}'].number_format = '0.00%'

        sheet[f'N{i+2}'] = resnet_m[i-1][0]
        sheet[f'N{i+2}'].number_format = '0.00%'
        sheet[f'O{i+2}'] = resnet_m[i-1][1]
        sheet[f'O{i+2}'].number_format = '0.00%'
        sheet[f'P{i+2}'] = resnet_m[i-1][2]
        sheet[f'P{i+2}'].number_format = '0.00%'
        sheet[f'Q{i+2}'] = resnet_m[i-1][3]
        sheet[f'Q{i+2}'].number_format = '0.00%'
    
    sheet[f'A15'] = 'Avg'
    sheet[f'A15'].font = bold_font
    sheet[f'A15'].alignment = center_alignment
    lis = ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q"]
    for ch in lis:
        sheet[f'{ch}15'] = f"=ROUND(AVERAGE({ch}3:{ch}14),4)"
        sheet[f'{ch}15'].number_format = '0.00%'    

    
    # Save the workbook
    workbook.save('model_performance.xlsx')

def get_acc_sen_spe_vits(prob_lab_paths):
    pred_probs = pd.read_csv(prob_lab_paths[0],header=None).values.flatten()
    true_labels = pd.read_csv(prob_lab_paths[1],header=None).values.flatten()
    pred_labels = (pred_probs >= 0.5).astype(int)
    cm = confusion_matrix(true_labels, pred_labels)
    roc_auc_score= metrics.roc_auc_score(true_labels,pred_probs)
    TN, FP, FN, TP = cm.ravel()
    acc = (TP+TN)/(TP+FP+FN+TN)
    sensitivity = TP / (TP + FN)
    Specificity = TN / (TN + FP)
    return (round(roc_auc_score,4),round(acc,4),round(sensitivity,4),round(Specificity,4))

def get_acc_sen_spe_resnet(pred_probs,true_labels):
    pred_labels = (pred_probs >= 0.5).astype(int)
    cm = confusion_matrix(true_labels, pred_labels)
    roc_auc_score= metrics.roc_auc_score(true_labels,pred_probs)
    TN, FP, FN, TP = cm.ravel()
    acc = (TP+TN)/(TP+FP+FN+TN)
    sensitivity = TP / (TP + FN)
    Specificity = TN / (TN + FP)
    return (round(roc_auc_score,4),round(acc,4),round(sensitivity,4),round(Specificity,4))

def check_ones(sublist):
    """
    Check if more than 50 values in the sublist are 0.
    :param sublist: List of 100 values containing 0s and 1s.
    :return: 0 if more than 50 values are 0, otherwise 1.
    """
    sublist = sublist.tolist()
    one_count = sublist.count(1)
    if one_count > 10:
        return 1
    else:
        return 0

def get_majority_vote_pred(pred_lab):
    new_pred = []
    chunk_size=20
    for i in range(0,len(pred_lab), chunk_size):
        chunk = pred_lab[i:i + chunk_size]
        result = check_ones(chunk)
        new_pred.append(result)
    return new_pred

def get_acc_sen_spe_llms(prob_lab_paths,bm_num):
    print(prob_lab_paths)
    if bm_num in [4,6,8,9,11,12]:
        pred_probs = pd.read_csv(prob_lab_paths[0]).values.flatten()
        true_labels = pd.read_csv(prob_lab_paths[1][1]).values.flatten()
    else:
        pred_probs = pd.read_csv(prob_lab_paths[0],header=None).values.flatten()
        true_labels = pd.read_csv(prob_lab_paths[1][1],header=None).values.flatten()
    majority_true_labels = pd.read_csv(prob_lab_paths[1][0],header=None).values.flatten()
    
    roc_auc_score= metrics.roc_auc_score(true_labels,pred_probs)
    print(f"len of majority true lables {len(majority_true_labels)} len of pred_probs {len(pred_probs)}, len of true label {len(true_labels)}")
    
    pred_labels = (pred_probs >= 0.5).astype(int)
    majority_pred = get_majority_vote_pred(pred_labels)
    print(f"len of true lables {len(majority_true_labels)} len of pred_probs {len(majority_pred)}")
    
    cm = confusion_matrix(majority_true_labels, majority_pred)
    TN, FP, FN, TP = cm.ravel()
    acc = (TP+TN)/(TP+FP+FN+TN)
    sensitivity = TP / (TP + FN)
    Specificity = TN / (TN + FP)
    return (round(roc_auc_score,4),round(acc,4),round(sensitivity,4),round(Specificity,4))

base_path = "./results/"
base_path_mit = base_path+"mit/"
base_path_espformer = base_path+"ESPFormer/"
base_path_resnet = base_path + "resnet/"
base_path_longformer = base_path+"longformer/"
base_path_mit_custom = base_path+"mit_custom/"
base_path_espformer_custom = base_path+"ESPFormer_custom/"
base_path_longformer_custom = base_path+"longformer_custom/"
MIT_metrics = []
ESPFormer_metrics = []
Longfor_metrics= []
resnet_metrics = []
sph_sop_list = [(1,2),(2,2),(5,2),(1,5),(2,5),(5,5),(1,15),(2,15),(5,15),(1,30),(2,30),(5,30)]
y_prob_valid_list = []
y_true_valid_list = []

for i in range(1,13):
    if i in range(1,10):
        new_i = "0"+str(i)
    j = i-1
    sph = sph_sop_list[j][1]
    sop = sph_sop_list[j][0]
    npy_fname = base_path_resnet + f'tuhszr_sngfld_unscld_unfilt_blcdet_srate256Hz_bmrk{i:02d}_sph{sph:02d}m_sop{sop:02d}m_seg05s_ovr00s_fold00_tuhstd_valid_output.npy'
    csv_fname = base_path_resnet + f'tuhszr_sngfld_unscld_unfilt_blcdet_srate256Hz_bmrk{i:02d}_sph{sph:02d}m_sop{sop:02d}m_seg05s_ovr00s_fold00_tuhstd_valid_labels.csv'
    
    y_prob_valid = np.load(npy_fname)
    y_true_valid = pd.read_csv(csv_fname,header=None).values
    resnet_metrics.append(get_acc_sen_spe_resnet(y_prob_valid,y_true_valid))

    
    file1_prob = f"OG_probablity_BM{new_i}.csv"
    file1_label = f"OG_True_label_BM{new_i}.csv"
    file1_label_LLM = f"Mjority_vote_True_label_BM{new_i}.csv"
    prob_file_name = f"BM{i}_probablity.csv"
    labels_file_name = f"BM{i}_labels.csv"
    labels_file_name_LLM = f"Majority_vote_BM{i}_labels.csv"
    
    MIT_metrics.append(get_acc_sen_spe_vits((load_csv_file(file1_prob,prob_file_name,base_path_mit_custom,base_path_mit),
                        load_csv_file(file1_label,labels_file_name,base_path_mit_custom,base_path_mit))))
    
    ESPFormer_metrics.append(get_acc_sen_spe_vits((load_csv_file(prob_file_name,prob_file_name,base_path_espformer_custom,base_path_espformer),
                        load_csv_file(labels_file_name,labels_file_name,base_path_espformer_custom,base_path_espformer))))
    
    Longfor_metrics.append(get_acc_sen_spe_llms((load_csv_file(file1_prob,prob_file_name,base_path_longformer_custom,base_path_longformer),
                         load_csv_file(file1_label_LLM,labels_file_name_LLM,base_path_longformer_custom,base_path_longformer,filename3=labels_file_name)),i))

#resnet remaining and putting them into excel is remaning
create_excel(MIT_metrics,Longfor_metrics,resnet_metrics,ESPFormer_metrics)